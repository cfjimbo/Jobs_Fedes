from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import filedialog
from io import BytesIO
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    return archivo

def leer_archivo_excel(nombre_archivo):
    try:
        df = pd.read_excel(nombre_archivo)
        return df
    except Exception as e:
        return None

def llenar_celdas_vacias_con_cero(df):
    try:
        df = df.fillna(0)
        return df
    except Exception as e:
        return None

def ajustar_numeros(df):
    try:
        for index, row in df.iterrows():
            if 'NUMERO' in df.columns:
                if isinstance(row['NUMERO'], int):
                    continue
                elif isinstance(row['NUMERO'], str) and row['NUMERO'].startswith('0') and len(row['NUMERO']) > 1:
                    df.at[index, 'NUMERO'] = int(row['NUMERO'][1:])
            elif 'Referencia' in df.columns:
                if isinstance(row['Referencia'], int):
                    continue
                elif isinstance(row['Referencia'], str) and row['Referencia'].startswith('0') and len(row['Referencia']) > 1:
                    df.at[index, 'Referencia'] = int(row['Referencia'][1:])
        return df
    except Exception as e:
        return None

def convertir_a_string(df):
    try:
        df = df.astype(str)
        return df
    except Exception as e:
        return None

def resaltar_coincidencias(df1, df2):
    try:
        filas_coincidentes = []
        filas_no_coincidentes_df1 = []
        filas_no_coincidentes_df2 = []
        
        for index1, row1 in df1.iterrows():
            coincidencia_encontrada = False
            for index2, row2 in df2.iterrows():
                if (row1['NUMERO'] == row2['Referencia'] and
                    row1['ACREDITA'] == row2['Creditos'] and
                    row1['DEBITA'] == row2['Debitos']):
                    fila_resaltada = {}
                    for columna in df1.columns:
                        fila_resaltada[columna] = row1[columna]
                    for columna in df2.columns:
                        fila_resaltada[columna] = row2[columna]
                    filas_coincidentes.append(fila_resaltada)
                    coincidencia_encontrada = True
                    break
            if not coincidencia_encontrada:
                fila_no_coincidente_df1 = {}
                for columna in df1.columns:
                    fila_no_coincidente_df1[columna] = row1[columna]
                filas_no_coincidentes_df1.append(fila_no_coincidente_df1)
        
        for index2, row2 in df2.iterrows():
            coincidencia_encontrada = False
            for index1, row1 in df1.iterrows():
                if (row1['NUMERO'] == row2['Referencia'] and
                    row1['ACREDITA'] == row2['Creditos'] and
                    row1['DEBITA'] == row2['Debitos']):
                    coincidencia_encontrada = True
                    break
            if not coincidencia_encontrada:
                fila_no_coincidente_df2 = {}
                for columna in df2.columns:
                    fila_no_coincidente_df2[columna] = row2[columna]
                filas_no_coincidentes_df2.append(fila_no_coincidente_df2)
        
        df_coincidentes = pd.DataFrame(filas_coincidentes)
        df_no_coincidentes_df1 = pd.DataFrame(filas_no_coincidentes_df1)
        df_no_coincidentes_df2 = pd.DataFrame(filas_no_coincidentes_df2)
        
        df_resultado = pd.concat([df_coincidentes, df_no_coincidentes_df1, df_no_coincidentes_df2], ignore_index=True)
        
        return df_resultado
    except Exception as e:
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/procesar', methods=['POST'])
def procesar():
    archivo1 = request.files['archivo1']
    archivo2 = request.files['archivo2']

    df1 = leer_archivo_excel(archivo1)
    df2 = leer_archivo_excel(archivo2)

    if df1 is not None and df2 is not None:
        df1 = llenar_celdas_vacias_con_cero(df1)
        df2 = llenar_celdas_vacias_con_cero(df2)
        
        df1 = ajustar_numeros(df1)
        df2 = ajustar_numeros(df2)

        df1 = convertir_a_string(df1)
        df2 = convertir_a_string(df2)

        df_resultado = resaltar_coincidencias(df1, df2)

        if df_resultado is not None:
            # Create a BytesIO object to save the result
            output = BytesIO()
            wb = Workbook()
            ws = wb.active

            for col_num, value in enumerate(df_resultado.columns, 1):
                ws.cell(row=1, column=col_num, value=value)

            for r_idx, row in enumerate(dataframe_to_rows(df_resultado, index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row_idx in range(2, len(df_resultado) + 2):
                for col_idx in range(1, len(df_resultado.columns) + 1):
                    if ws.cell(row=row_idx, column=col_idx).value:
                        ws.cell(row=row_idx, column=col_idx).fill = amarillo

            wb.save(output)
            output.seek(0)

            # Send the file as response to the client
            return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name='resultado.xlsx')
        
        else:
            return "No se pudieron resaltar las coincidencias correctamente."

if __name__ == '__main__':
    app.run(debug=True)
